from openpyxl import Workbook
from openpyxl import load_workbook

sourcePath='/Users/naanast/Desktop/Project/data-for-pythin.xlsx'
destPath='/Users/naanast/Desktop/Project/result.xlsx'
destTxtPath="/Users/naanast/Desktop/Project/result.txt"

wb1 = load_workbook(sourcePath)  # read data from source workbook
wb2 = Workbook() #open dest workbook

#open sheet
w1 =wb1.active
w2 = wb2.active
#name second column of new sheet
w2.cell(row=1, column=2).value="Total"
#copy first column from table
for x in range(1,w1.max_row+1):
   w2.cell(row=x,column= 1).value=w1.cell(row =x,column=1).value


for x in range(2,w1.max_row+1):
    total=0.0
    #loop in data table
    for y in range(2,w1.max_column+1):
        #loop to calc sum of data from row x
        if str(w1.cell(row=x,column=y).value)!="null":
             total = total+ float(w1.cell(row=x,column=y).value)

    w2.cell(row=x,column=2).value=total  #save total from  row x data
#save xlsx doc
wb2.save(destPath)


#save txt file
f = open(destTxtPath, "a")
for x in range (1,w2.max_row+1):
    for y in range(1,w2.max_column+1):
        f.write(str(w2.cell(row=x,column=y).value)+"\t")
    f.write("\n")
f.close()# close txt file
