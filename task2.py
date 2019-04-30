from openpyxl import Workbook
from openpyxl import load_workbook
import csv
sourcePath='/Users/naanast/Desktop/Project/task3/data/Source.xlsx'

destPath="/Users/naanast/Desktop/Project/task3/data/Q+report.xlsx"

compPath="/Users/naanast/Desktop/Project/task3/data/qplus-quality-report.csv"


wb1 = load_workbook(sourcePath)  # read data from source in a workbook
wb2 = Workbook() #open dest workbook

#open sheet
w1 =wb1.active
w2 = wb2.active


#create dest head table
w2.cell(row=1,column=1).value="Building"
w2.cell(row=1,column=2).value="Manager"
w2.cell(row=1,column=3).value="Case"
w2.cell(row=1,column=4).value="Ratings"
w2.cell(row=1,column=5).value="Reviewed"
w2.cell(row=1,column=6).value="Profile"
w2.cell(row=1,column=7).value="Amount"



for z in range (2,w1.max_row+1):


        count = 0
        case_id =str(w1.cell(row=z,column=18).value)

        with open(compPath) as csvfile:
            readCSV = csv.reader(csvfile,delimiter=',')

            for row in readCSV:


                if case_id == row[5]:
                    manager = str(row[3])
                    count=count+1


            w2.cell(row=z,column=1).value= w1.cell(row=z,column=7).value
            w2.cell(row=z,column=2).value=w1.cell(row=z,column=31).value
            w2.cell(row=z,column=3).value=str(case_id)
            w2.cell(row=z,column=4).value=w1.cell(row=z,column=38).value
            if count !=0:
                reviewed=1
            else:
                reviewed =0
            w2.cell(row=z,column=5).value=reviewed
            w2.cell(row=z,column=6).value=w1.cell(row=z,column=36).value
            w2.cell(row=z,column=7).value=str(count)







#save xlsx doc
wb2.save(destPath)














"""
#copy first column from table
for x in range(1,w1.max_row+1):
   w2.cell(row=x,column= 1).value=w1.cell(row =x,column=1).value


for x in range(2,w1.max_row+1):
    total=0.0
    #loop in data table
    for y in range(2,w1.max_column+1):
        #loop to calc sum of data from row x
        if str(w1.cell(row=x,column=y).value)!="null":
             total = total+ float(w1.cell(row=x,column=y).value)

    w2.cell(row=x,column=2).value=total  #save total from  row x data
    @@@@@@@@@@@@@@@@@@@@@@@@@@@
w2.cell(row=z,column=1).value= w1.cell(row=x,column=7)
w2.cell(row=z,column=2).value=str(row[3])
w2.cell(row=z,column=3).value=str(case_id)
w2.cell(row=z,column=4).value=w1.cell(row=x,column=38)
if count !=0:
    reviewed=1
else:
    rewiewed =0
w2.cell(row=z,column=5).value=str(reviewed)
w2.cell(row=z,column=6).value=w1.cell(row=x,column=36)
w2.cell(row=z,column=7).value=str(count)






"""


